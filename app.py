import os, re, io, zipfile, datetime
from collections import defaultdict
from flask import Flask, render_template, request, send_file, jsonify
import openpyxl

CODIGOS_MENUDENCIAS = {'FD606006', 'FD220507', 'FD219009'}

app = Flask(__name__, template_folder=".")
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
PLANT_DIR = BASE_DIR

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/tiene_menudencias', methods=['POST'])
def tiene_menudencias():
    try:
        piqueo_f = request.files.get('piqueo')
        if not piqueo_f:
            return jsonify({'tiene': False})
        piqueo = leer_piqueo(piqueo_f)
        codigos = set(piqueo['cortes'].keys())
        tiene = bool(codigos & CODIGOS_MENUDENCIAS)
        return jsonify({'tiene': tiene})
    except Exception as e:
        return jsonify({'tiene': False, 'error': str(e)})


@app.route('/generar', methods=['POST'])
def generar():
    try:
        piqueo_f  = request.files.get('piqueo')
        reporte_f = request.files.get('reporte')
        remito_f  = request.files.get('remito')
        shipment  = request.form.get('shipment_no', '').strip()
        destino   = request.form.get('destino', '').strip()
        tipo      = request.form.get('tipo', 'carne').strip()

        errores = []
        if not piqueo_f:  errores.append("Falta el Piqueo (.xlsx)")
        if not reporte_f: errores.append("Falta el Reporte DOC (.xlsx)")
        if not remito_f:  errores.append("Falta el Remito (PDF que contiene REMITO en el nombre)")
        if not shipment:  errores.append("Falta el numero de shipment")
        if not destino:   errores.append("Selecciona el destino")
        if errores:
            return jsonify({'ok': False, 'errores': errores}), 400

        piqueo  = leer_piqueo(piqueo_f)
        reporte = leer_reporte(reporte_f, shipment)
        remito  = leer_remito(remito_f)

        plantilla = os.path.join(PLANT_DIR, f'PLANTILLA_{destino}.docx')
        if not os.path.exists(plantilla):
            return jsonify({'ok': False, 'errores': [
                f'Plantilla para "{destino}" no encontrada en el servidor.'
            ]}), 500

        with open(plantilla, 'rb') as f:
            docx_bytes = f.read()

        resultado, alertas = generar_sanitario(docx_bytes, piqueo, reporte, remito, destino, tipo)

        sufijo = f'_{tipo}' if destino in ('grecia_directo', 'grecia_rotterdam') else ''
        nombre = f"Sanitario_{destino}_{shipment}{sufijo}.docx"
        resp = send_file(
            io.BytesIO(resultado), as_attachment=True,
            download_name=nombre,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        if alertas:
            resp.headers['X-Alertas'] = ' | '.join(alertas)
        return resp

    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'errores': [str(e), traceback.format_exc()]}), 500


def leer_piqueo(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    contenedor_piq = None
    for row in list(ws.iter_rows(values_only=True))[:5]:
        if row and len(row) > 5 and 'CONT' in str(row[3] or ''):
            contenedor_piq = str(row[5] or '').replace(' ', '')

    cortes = defaultdict(lambda: {'nombre':'','cajas':0,'neto':0.0,'prod_desde':None,'prod_hasta':None})
    prod_desde_all = []
    prod_hasta_all = []
    en_datos = False
    for row in ws.iter_rows(values_only=True):
        if not any(row): continue
        if str(row[1] or '').strip() == 'Cod Prod':
            en_datos = True; continue
        if not en_datos: continue
        if len(row) < 7: continue
        cod, nom, cajas = row[1], row[2], row[3]
        kg_neto = row[5]   # col 5 = Peso (neto), col 4 = Unidades
        pd_ = row[6]       # col 6 = Fecha P (produccion)
        if not cod or 'Total' in str(row[0] or ''): continue
        g = cortes[str(cod)]
        g['nombre'] = nom or g['nombre']
        g['cajas'] += int(cajas) if cajas else 0
        g['neto']  += float(kg_neto) if kg_neto else 0.0
        if isinstance(pd_, datetime.datetime):
            g['prod_desde'] = min(g['prod_desde'], pd_) if g['prod_desde'] else pd_
            g['prod_hasta'] = max(g['prod_hasta'], pd_) if g['prod_hasta'] else pd_
            prod_desde_all.append(pd_)
            prod_hasta_all.append(pd_)

    faena = defaultdict(lambda: {'desde':None,'hasta':None})
    faena_all = []
    if 'Pegar Datos' in wb.sheetnames:
        for row in list(wb['Pegar Datos'].iter_rows(values_only=True))[1:]:
            if not row[1]: continue
            cod, ff = str(row[1]), row[7]
            if isinstance(ff, datetime.datetime):
                faena[cod]['desde'] = min(faena[cod]['desde'], ff) if faena[cod]['desde'] else ff
                faena[cod]['hasta'] = max(faena[cod]['hasta'], ff) if faena[cod]['hasta'] else ff
                faena_all.append(ff)

    return {
        'cortes': dict(cortes),
        'faena': dict(faena),
        'faena_desde': min(faena_all).strftime('%d/%m/%Y') if faena_all else None,
        'faena_hasta': max(faena_all).strftime('%d/%m/%Y') if faena_all else None,
        'prod_desde':  min(prod_desde_all).strftime('%d/%m/%Y') if prod_desde_all else None,
        'prod_hasta':  max(prod_hasta_all).strftime('%d/%m/%Y') if prod_hasta_all else None,
        'contenedor': contenedor_piq,
    }


def leer_reporte(file, shipment):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return {}

    # Buscar la fila de headers (puede estar en fila 0 o fila 1 segun el reporte)
    hdr = None
    hdr_idx = 0
    for i, row in enumerate(rows[:5]):
        if row and any(str(v or '').strip() in ('Shipment No', 'Shipment') for v in row):
            hdr = row
            hdr_idx = i
            break
    if hdr is None:
        hdr = rows[0]
        hdr_idx = 0

    def col(keys):
        for i, h in enumerate(hdr):
            if h and any(k.lower() in str(h).lower() for k in keys): return i
        return None

    c_ship = col(['Shipment No', 'Shipment'])
    c_ves  = col(['Vessel', 'Aircraft'])
    c_bl   = col(['B/L', 'BL', 'AWB'])
    c_cont = col(['Container'])
    c_dep  = col(['Departure'])
    c_desc = col(['Description'])
    c_code = col(['Code'])

    d = {'vessel':None,'bl':None,'contenedor':None,'fecha_salida':None,'descripciones':{}}

    ship_base = shipment.split('-')[0]
    for row in rows[hdr_idx+1:]:
        if not row or c_ship is None: continue
        sv = str(row[c_ship] or '').strip()
        if sv != shipment and not sv.startswith(ship_base): continue
        def get(c): return str(row[c]).strip() if c is not None and row[c] else None
        if not d['vessel']     and get(c_ves):  d['vessel']  = get(c_ves)
        if not d['bl']         and get(c_bl):   d['bl']      = get(c_bl)
        if not d['contenedor'] and get(c_cont): d['contenedor'] = get(c_cont)
        if not d['fecha_salida'] and c_dep is not None and row[c_dep]:
            v = row[c_dep]
            d['fecha_salida'] = v.strftime('%d/%m/%Y') if isinstance(v, datetime.datetime) else str(v).strip()
        if c_code and c_desc and row[c_code] and row[c_desc]:
            d['descripciones'][str(row[c_code]).strip()] = str(row[c_desc]).strip()
    return d


def leer_remito(file):
    texto = ''
    try:
        import fitz
        pdf_bytes = file.read()
        doc = fitz.open(stream=pdf_bytes, filetype='pdf')
        for page in doc:
            texto += page.get_text() + '\n'
        doc.close()
    except Exception:
        return {}

    def buscar(patron, grupo=1):
        m = re.search(patron, texto, re.IGNORECASE)
        return m.group(grupo).strip() if m else None

    def limpiar_num(s):
        if not s: return None
        s = s.replace('.', '').replace(',', '.')
        try: return str(round(float(s), 2))
        except: return s

    cont = buscar(r'CONTAINER[:\s]+([A-Z]{4}\d{6,7}-?\d?)')
    if cont: cont = cont.replace(' ', '')

    return {
        'texto':           texto,
        'contenedor':      cont,
        'precinto_senasa': buscar(r'P\.S\.[:\s]+(\d{6,7})'),
        'precinto_afip':   buscar(r'P\.A\.[:\s]+([A-Z]{2,3}\d{4,8})'),
        'contramarca':     buscar(r'CONTRAMARCA[:\s]+(\d{9,12})'),
        'peso_neto':       limpiar_num(buscar(r'PESO NETO TOTAL[:\s]+([\d\.,]+)')),
        'peso_bruto':      limpiar_num(buscar(r'PESO BRUTO TOTAL[:\s]+([\d\.,]+)')),
        'total_cajas':     int(buscar(r'Total General\s+([\d\.]+)').replace('.','')) if buscar(r'Total General\s+([\d\.]+)') else None,
        'vessel':          buscar(r'Buque/Aerol[ií]nea[:\s]+([A-Z][A-Z ]+?)(?:\n|DESTINO|$)'),
        'fecha_emision':   buscar(r'FECHA[:\s]+(\d{2}\.\d{2}\.\d{4})'),
    }


def consolidar(piqueo, reporte, remito):
    alertas = []
    vessel = reporte.get('vessel') or remito.get('vessel') or 'FALTA_VESSEL'
    bl = reporte.get('bl') or ''
    if not bl:
        bl = 'FALTA_BL'
        alertas.append('BL VACIO - completar manualmente antes de presentar')
    cont   = remito.get('contenedor')      or piqueo.get('contenedor') or 'FALTA_CONTENEDOR'
    prec_s = remito.get('precinto_senasa') or 'FALTA_PRECINTO_SENASA'
    prec_a = remito.get('precinto_afip')   or 'FALTA_PRECINTO_AFIP'
    contra = remito.get('contramarca')     or 'FALTA_CONTRAMARCA'
    f_sal  = reporte.get('fecha_salida')   or 'FALTA_FECHA_SALIDA'
    if f_sal == 'FALTA_FECHA_SALIDA':
        alertas.append('FECHA DE SALIDA no esta en el Reporte DOC - completar manualmente')
    f_emi_raw = remito.get('fecha_emision') or ''
    f_emi = f_emi_raw.replace('.', '/') if f_emi_raw and '.' in f_emi_raw else datetime.datetime.now().strftime('%d/%m/%Y')
    faena_d = piqueo.get('faena_desde') or 'FALTA'
    faena_h = piqueo.get('faena_hasta') or 'FALTA'
    prod_d  = piqueo.get('prod_desde')  or 'FALTA'
    prod_h  = piqueo.get('prod_hasta')  or 'FALTA'
    neto    = remito.get('peso_neto')   or '0'
    bruto   = remito.get('peso_bruto')  or '0'
    cajas   = remito.get('total_cajas') or sum(c['cajas'] for c in piqueo['cortes'].values())
    return {
        'vessel': vessel, 'bl': bl, 'contenedor': cont,
        'precinto_senasa': prec_s, 'precinto_afip': prec_a,
        'contramarca': contra,
        'fecha_salida': f_sal, 'fecha_emision': f_emi,
        'faena_desde': faena_d, 'faena_hasta': faena_h,
        'prod_desde': prod_d, 'prod_hasta': prod_h,
        'peso_neto': neto, 'peso_bruto': bruto, 'total_cajas': cajas,
    }, alertas


def peso_fmt_std(neto, bruto, sep='  / '):
    try: return f"{float(neto):.2f} KGS{sep}{float(bruto):.2f} KGS"
    except: return f"{neto} KGS{sep}{bruto} KGS"

def get_trs(xml):
    return list(re.finditer(r'<w:tr\s', xml))

def get_fila_xml(xml, trs, idx):
    ini = trs[idx].start()
    fin = trs[idx+1].start() if idx+1 < len(trs) else len(xml)
    return xml[ini:fin], ini, fin

def construir_fila_corte(fila_modelo, cajas, descripcion, neto):
    nueva = fila_modelo
    def reemplazar_celda_texto(xml_fila, celda_idx, nuevo_texto):
        celda_starts = [m.start() for m in re.finditer(r'<w:tc>', xml_fila)]
        celda_ends   = [m.start() for m in re.finditer(r'</w:tc>', xml_fila)]
        if celda_idx >= len(celda_starts): return xml_fila
        bloque = xml_fila[celda_starts[celda_idx]:celda_ends[celda_idx]]
        textos_en_celda = re.findall(r'<w:t[^>]*>[^<]*</w:t>', bloque)
        if not textos_en_celda: return xml_fila
        primer = textos_en_celda[0]
        tag_open = re.match(r'<w:t[^>]*>', primer).group()
        nuevo_run = f'{tag_open}{nuevo_texto}</w:t>'
        nuevo_bloque = bloque.replace(primer, nuevo_run, 1)
        for t in textos_en_celda[1:]:
            tag_open2 = re.match(r'<w:t[^>]*>', t).group()
            nuevo_bloque = nuevo_bloque.replace(t, f'{tag_open2}</w:t>', 1)
        return xml_fila[:celda_starts[celda_idx]] + nuevo_bloque + xml_fila[celda_ends[celda_idx]:]
    nueva = reemplazar_celda_texto(nueva, 0, str(cajas))
    nueva = reemplazar_celda_texto(nueva, 1, descripcion)
    nueva = reemplazar_celda_texto(nueva, 2, str(neto))
    return nueva

def reemplazar_fila_total(fila_total_xml, total_cajas, total_neto):
    nums = re.findall(r'<w:t[^>]*>(\d{2,6})</w:t>', fila_total_xml)
    nueva = fila_total_xml
    if len(nums) >= 2:
        nueva = nueva.replace(f'<w:t>{nums[0]}</w:t>', f'<w:t>{total_cajas}</w:t>', 1)
        nueva = nueva.replace(f'<w:t>{nums[1]}</w:t>', f'<w:t>{total_neto}</w:t>', 1)
    return nueva

def reemplazar_bloque_cortes(xml, trs, primera_fila_idx, total_idx, cortes_nuevos, total_cajas, total_neto):
    fila_modelo, ini_mod, _ = get_fila_xml(xml, trs, primera_fila_idx)
    fila_total, ini_tot, fin_tot = get_fila_xml(xml, trs, total_idx)
    nuevas_filas = ''
    for corte in cortes_nuevos:
        nuevas_filas += construir_fila_corte(fila_modelo, corte['cajas'], corte['descripcion'], corte['neto'])
    nueva_total = reemplazar_fila_total(fila_total, total_cajas, total_neto)
    return xml[:ini_mod] + nuevas_filas + nueva_total + xml[fin_tot:]

def _set_condicion_transporte(xml, es_congelado):
    m = re.search(r'De refrigeraci', xml, re.IGNORECASE)
    if not m:
        return xml
    tr_start = xml.rfind('<w:tr ', 0, m.start())
    if tr_start == -1:
        tr_start = xml.rfind('<w:tr>', 0, m.start())
    tr_end = xml.find('</w:tr>', m.start()) + 7
    fila_original = xml[tr_start:tr_end]
    celda_matches = list(re.finditer(r'<w:tc>', fila_original))
    if len(celda_matches) < 5:
        return xml
    def get_celda(fila, idx):
        start = celda_matches[idx].start()
        end = fila.find('</w:tc>', start) + 7
        return fila[start:end]
    celda2 = get_celda(fila_original, 2)
    celda4 = get_celda(fila_original, 4)
    def celda_con_x(celda_xml):
        celda_xml = re.sub(r'<w:r\b[^>]*>(?:(?!</w:r>).)*<w:t[^>]*>\s*[xX]\s*</w:t>(?:(?!</w:r>).)*</w:r>',
                           '', celda_xml, flags=re.DOTALL)
        celda_xml = re.sub(r'(</w:pPr>)(</w:p>)',
                           r'\1<w:r><w:rPr><w:rFonts w:ascii="Calibri" w:hAnsi="Calibri"/>'
                           r'<w:sz w:val="16"/><w:szCs w:val="16"/></w:rPr>'
                           r'<w:t>X</w:t></w:r>\2',
                           celda_xml, count=1)
        return celda_xml
    def celda_sin_x(celda_xml):
        return re.sub(r'<w:r\b[^>]*>(?:(?!</w:r>).)*<w:t[^>]*>\s*[xX]\s*</w:t>(?:(?!</w:r>).)*</w:r>',
                      '', celda_xml, flags=re.DOTALL)
    if es_congelado:
        nueva_celda2 = celda_sin_x(celda2)
        nueva_celda4 = celda_con_x(celda4)
    else:
        nueva_celda2 = celda_con_x(celda2)
        nueva_celda4 = celda_sin_x(celda4)
    fila_nueva = fila_original.replace(celda2, nueva_celda2, 1).replace(celda4, nueva_celda4, 1)
    return xml[:tr_start] + fila_nueva + xml[tr_end:]


def _es_congelado(piqueo):
    for cod in piqueo['cortes']:
        return str(cod).upper().startswith('F')
    return False


def _reemplazar_fecha_roja_i14(xml, nueva_fecha):
    patron = (r'(<w:color w:val="FF0000"/>.*?<w:lang w:val="es-AR"/>.*?</w:rPr>\s*<w:t>)'
              r'(\d{2}/\d{2}/\d{4})(</w:t>)')
    m = re.search(patron, xml, re.DOTALL)
    if m:
        xml = xml[:m.start()] + m.group(1) + nueva_fecha + m.group(3) + xml[m.end():]
    return xml

def _listar_fechas_faena(piqueo):
    fechas = set()
    for datos in piqueo['faena'].values():
        for k in ('desde', 'hasta'):
            if datos.get(k):
                fechas.add(datos[k].strftime('%d/%m/%Y'))
    return sorted(fechas)

def _listar_fechas_produccion(piqueo):
    fechas = set()
    for datos in piqueo['cortes'].values():
        for k in ('prod_desde', 'prod_hasta'):
            if datos.get(k):
                fechas.add(datos[k].strftime('%d/%m/%Y'))
    return sorted(fechas)

def _extraer_tropas(texto):
    m = re.search(r'TROPAS?[:\s]+([0-9\-/]+(?:[,\s]+[0-9\-/]+)*)', texto, re.IGNORECASE)
    if m:
        nums = re.findall(r'\d+', m.group(1))
        return '-'.join(nums)
    return ''


MAPA_CORTES = {
    'CORAZON DE CUADRIL': {
        'alemania_directo':   'CORAZON DE CUADRIL / HEART OF RUMP',
        'alemania_rotterdam': 'CORAZON DE CUADRIL / HEART OF RUMP',
        'espana':             'CORAZON DE CUADRIL / HEART OF RUMP',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'CORAZON DE CUADRIL / HEART OF RUMP',
    },
    'LOMO SC': {
        'alemania_directo':   'LOMO SIN CORDON {lbs} / TENDERLOIN CHAIN OFF {lbs}',
        'alemania_rotterdam': 'LOMO SIN CORDON {lbs} / TENDERLOIN CHAIN OFF {lbs}',
        'espana':             'LOMO SIN CORDON {lbs} / TENDERLOIN CHAIN OFF {lbs}',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'LOMO SIN CORDON {lbs} / TENDERLOIN CHAIN OFF {lbs}',
    },
    'BIFE ANCHO': {
        'alemania_directo':   'BIFE ANCHO SC / RIBEYE ROLL SC',
        'alemania_rotterdam': 'BIFE ANCHO SIN TAPA/ RIB EYE',
        'espana':             'BIFE ANCHO SIN TAPA / RIBEYE',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'BIFE ANCHO SIN TAPA / RIB EYE',
    },
    'BIFE ANGOSTO': {
        'alemania_directo':   'BIFE ANGOSTO / STRIP LOIN',
        'alemania_rotterdam': 'BIFE ANGOSTO/ STRIPLOIN',
        'espana':             'BIFE ANGOSTO / STRIPLOIN',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'BIFE ANGOSTO / STRIPLOIN',
    },
    'BOLA DE LOMO': {
        'alemania_directo':   'BOLA DE LOMO / KNUCKLE',
        'alemania_rotterdam': 'BOLA DE LOMO/ KNUCKLE',
        'espana':             'BOLA DE LOMO / KNUCKLE',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'BOLA DE LOMO / KNUCKLE',
    },
    'CUADRADA': {
        'alemania_directo':   'CUADRADA/ FLAT',
        'alemania_rotterdam': 'CUADRADA / FLAT',
        'espana':             'CUADRADA / FLAT',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'CUADRADA / FLAT',
    },
    'NALGA': {
        'alemania_directo':   'NALGA SIN TAPA / TOPSIDE CAP OFF',
        'alemania_rotterdam': 'NALGA SIN TAPA/ TOPSIDE CAP OFF',
        'espana':             'NALGA SIN TAPA / TOPSIDE CAP OFF',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'NALGA ST / TOPSIDE CAP OFF',
    },
    'PECETO': {
        'alemania_directo':   'PECETO / EYEROUND',
        'alemania_rotterdam': 'PECETO / EYEROUND',
        'espana':             'PECETO / EYEROUND',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'PECETO / EYEROUND',
    },
    'COLITA DE CUADRIL': {
        'alemania_directo':   'COLITA DE CUADRIL / TAIL OF RUMP',
        'alemania_rotterdam': 'COLITA DE CUADRIL / TAIL OF RUMP',
        'espana':             'COLITA DE CUADRIL / TAIL OF RUMP',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'COLITA DE CUADRIL / TAIL OF RUMP',
    },
    'TAPA DE CUADRIL': {
        'alemania_directo':   'TAPA DE CUADRIL / CAP OF RUMP',
        'alemania_rotterdam': 'TAPA DE CUADRIL/ CAP OF RUMP',
        'espana':             'TAPA DE CUADRIL {kg} / TAIL OF RUMP',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'TAPA DE CUADRIL / RUMP CAP',
    },
    'AGUJA': {
        'alemania_directo':   'AGUJA / CHUCK',
        'alemania_rotterdam': 'AGUJA / CHUCK',
        'espana':             'AGUJA / CHUCK',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'AGUJA/ CHUCK ROLL',
    },
    'MARUCHA': {
        'alemania_directo':   'MARUCHA / OYSTER BLADE',
        'alemania_rotterdam': 'MARUCHA/ OYSTER BLADE',
        'espana':             'MARUCHA / OYSTER BLADE',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'MARUCHA / OYSTER BLADE',
    },
    'CORAZON DE PALETA': {
        'alemania_directo':   'CORAZON DE PALETA / SHOULDER CLOD HEART',
        'alemania_rotterdam': 'CORAZON DE PALETA/ SHOULDER CLOD HEART',
        'espana':             'CORAZON DE PALETA / SHOULDER CLOD HEART',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'CORAZON DE PALETA / SHOULDER CLOD HEART',
    },
    'CHINGOLO': {
        'alemania_directo':   'CHINGOLO / CHUCK TENDER',
        'alemania_rotterdam': 'CHINGOLO / CHUCK TENDER',
        'espana':             'CHINGOLO / CHUCK TENDER',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'CHINGOLO / CHUCK TENDER',
    },
    'BIFE DE VACIO': {
        'alemania_directo':   'BIFE DE VACIO / FLANK',
        'alemania_rotterdam': 'BIFE DE VACIO GRANDE/ FLAP MEAT',
        'espana':             'BIFE DE VACIO / FLANK',
        'grecia_directo':     'TROZOS DE CARNE VACUNA CONGELADA SIN HUESO / FROZEN BEEF PIECES IN BLOCK FOR INDUSTRIAL USE',
        'grecia_rotterdam':   'BIFE DE VACIO GRANDE / FLAP MEAT',
    },
}

def nombre_oficial(nombre_piqueo, destino):
    n = nombre_piqueo.upper()
    for clave, mapa in MAPA_CORTES.items():
        if clave in n:
            nombre = mapa.get(destino, nombre_piqueo)
            if '{lbs}' in nombre:
                lbs = ''
                for tag in ['3/4 LBS','4/5 LBS','+5 LBS']:
                    if tag in n: lbs = tag; break
                nombre = nombre.replace('{lbs}', lbs).strip()
            if '{kg}' in nombre:
                kg = '-1,3 KG' if '-1' in n or 'CHICO' in n else '+1,3 KG'
                nombre = nombre.replace('{kg}', kg)
            return nombre
    return nombre_piqueo

def listar_cortes_para_anexo(piqueo, destino):
    """Una fila por codigo, sin agrupar."""
    resultado = []
    for cod, datos in piqueo['cortes'].items():
        nombre = nombre_oficial(datos['nombre'], destino)
        resultado.append({
            'descripcion': nombre,
            'cajas': datos['cajas'],
            'neto': int(round(datos['neto'])),
        })
    return resultado


def _filtrar_piqueo_por_tipo(piqueo, tipo):
    import copy
    p = copy.deepcopy(piqueo)
    if tipo == 'menudencias':
        p['cortes'] = {k: v for k, v in p['cortes'].items() if k in CODIGOS_MENUDENCIAS}
    else:
        p['cortes'] = {k: v for k, v in p['cortes'].items() if k not in CODIGOS_MENUDENCIAS}
    prod_desde_list = [v['prod_desde'] for v in p['cortes'].values() if v.get('prod_desde')]
    prod_hasta_list = [v['prod_hasta'] for v in p['cortes'].values() if v.get('prod_hasta')]
    p['prod_desde'] = min(prod_desde_list).strftime('%d/%m/%Y') if prod_desde_list else None
    p['prod_hasta'] = max(prod_hasta_list).strftime('%d/%m/%Y') if prod_hasta_list else None
    codigos = set(p['cortes'].keys())
    faena_filtrada = {k: v for k, v in p['faena'].items() if k in codigos}
    faena_all = []
    for v in faena_filtrada.values():
        if v.get('desde'): faena_all.append(v['desde'])
        if v.get('hasta'): faena_all.append(v['hasta'])
    p['faena_desde'] = min(faena_all).strftime('%d/%m/%Y') if faena_all else p.get('faena_desde')
    p['faena_hasta'] = max(faena_all).strftime('%d/%m/%Y') if faena_all else p.get('faena_hasta')
    return p


def generar_sanitario(docx_bytes, piqueo, reporte, remito, destino, tipo='carne'):
    if destino in ('grecia_directo', 'grecia_rotterdam'):
        piqueo = _filtrar_piqueo_por_tipo(piqueo, tipo)
    datos, alertas = consolidar(piqueo, reporte, remito)
    with zipfile.ZipFile(io.BytesIO(docx_bytes), 'r') as z:
        archivos = {n: z.read(n) for n in z.namelist()}
    xml = archivos['word/document.xml'].decode('utf-8')
    if destino == 'alemania_directo':
        xml, al = _gen_alemania_directo(xml, datos, piqueo)
    elif destino == 'alemania_rotterdam':
        xml, al = _gen_alemania_rotterdam(xml, datos, piqueo)
    elif destino == 'espana':
        xml, al = _gen_espana(xml, datos, piqueo)
    elif destino == 'grecia_directo':
        xml, al = _gen_grecia_directo(xml, datos, piqueo)
    elif destino == 'grecia_rotterdam':
        xml, al = _gen_grecia_rotterdam(xml, datos, piqueo)
    else:
        al = [f'Destino {destino} no implementado']
    alertas.extend(al)
    archivos['word/document.xml'] = xml.encode('utf-8')
    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as z:
        for n, d in archivos.items():
            z.writestr(n, d)
    out.seek(0)
    return out.read(), alertas


def _gen_alemania_directo(xml, d, piqueo):
    alertas = []
    xml = xml.replace('MAERSK LAGUNA', d['vessel'])
    xml = xml.replace('CGMU569012-8', d['contenedor'])
    xml = xml.replace('BAH66457', d['precinto_afip'])
    xml = xml.replace('BUA0364222', d['bl'])
    xml = xml.replace('4501986641', d['contramarca'])
    xml = xml.replace('21009.00 KGS  / 22244.70 KGS', peso_fmt_std(d['peso_neto'], d['peso_bruto']))
    xml = re.sub(r'(<w:t>)997(</w:t>)', lambda m: m.group(1)+str(d['total_cajas'])+m.group(2), xml, count=1)
    xml = _reemplazar_fecha_roja_i14(xml, d['fecha_salida'])
    xml = xml.replace('>08/04/2026<', f">{d['prod_desde']}<")
    xml = xml.replace('>18/04/2026<', f">{d['prod_hasta']}<")
    xml = xml.replace('31/03/2026', d['faena_desde'])
    xml = xml.replace('16/04/2026', d['faena_hasta'])
    xml = xml.replace('26/04/2026', d['fecha_emision'])
    xml = re.sub(r'(<w:t>)997(</w:t>)', lambda m: m.group(1)+str(d['total_cajas'])+m.group(2), xml)
    xml = re.sub(r'(<w:t>)21009(</w:t>)', lambda m: m.group(1)+str(int(float(d['peso_neto'] or 0)))+m.group(2), xml)
    cortes = listar_cortes_para_anexo(piqueo, 'alemania_directo')
    if cortes:
        trs = get_trs(xml)
        xml = reemplazar_bloque_cortes(xml, trs, 38, 46, cortes, d['total_cajas'], int(float(d['peso_neto'] or 0)))
    xml = _set_condicion_transporte(xml, _es_congelado(piqueo))
    return xml, alertas

def _gen_alemania_rotterdam(xml, d, piqueo):
    alertas = []
    xml = xml.replace('MAERSK LAGUNA', d['vessel'])
    xml = xml.replace('MNBU054740-8', d['contenedor'])
    xml = xml.replace('BAH66118', d['precinto_afip'])
    xml = xml.replace('BL Nº 265885386', f"BL Nº {d['bl']}")
    xml = xml.replace('K 26 11', d['contramarca'])
    xml = xml.replace('20471.00 KGS / 21808.02', f"{float(d['peso_neto']):.2f} KGS / {float(d['peso_bruto']):.2f}")
    xml = _reemplazar_fecha_roja_i14(xml, d['fecha_salida'])
    xml = xml.replace('03/02/2026', d['prod_desde'])
    xml = xml.replace('12/02/2026', d['prod_hasta'])
    xml = xml.replace('30/01/2026', d['faena_desde'])
    xml = xml.replace('10/02/2026', d['faena_hasta'])
    xml = xml.replace('23/02/2026', d['fecha_emision'])
    xml = re.sub(r'(<w:t>)1071(</w:t>)', lambda m: m.group(1)+str(d['total_cajas'])+m.group(2), xml)
    xml = re.sub(r'(<w:t>)20471(</w:t>)', lambda m: m.group(1)+str(int(float(d['peso_neto'] or 0)))+m.group(2), xml)
    cortes = listar_cortes_para_anexo(piqueo, 'alemania_rotterdam')
    if cortes:
        trs = get_trs(xml)
        xml = reemplazar_bloque_cortes(xml, trs, 39, 73, cortes, d['total_cajas'], int(float(d['peso_neto'] or 0)))
    xml = _set_condicion_transporte(xml, _es_congelado(piqueo))
    return xml, alertas

def _gen_espana(xml, d, piqueo):
    alertas = []
    xml = xml.replace('MAERSK LANCO', d['vessel'])
    xml = xml.replace('MNBU3683558', d['contenedor'])
    xml = xml.replace('BAH66449', d['precinto_afip'])
    xml = xml.replace('269379303', d['bl'])
    xml = re.sub(r'(<w:t[^>]*>)13290\.00\s*(</w:t>)',
                 lambda m: m.group(1)+f"{float(d['peso_neto']):.2f} "+m.group(2),
                 xml, count=1)
    xml = re.sub(r'(<w:t[^>]*>)14231\.27\s*(</w:t>)',
                 lambda m: m.group(1)+f"{float(d['peso_bruto']):.2f} "+m.group(2),
                 xml, count=1)
    xml = _reemplazar_fecha_roja_i14(xml, d['fecha_salida'])
    xml = xml.replace('22/03/2026', d['prod_desde'] or '22/03/2026')
    xml = xml.replace('13/04/2026', d['prod_hasta'] or '13/04/2026', 1)  # solo primera ocurrencia (I.27), el lote se maneja aparte
    xml = xml.replace('17/03/2026', d['faena_desde'])
    xml = xml.replace('09/04/2026', d['faena_hasta'])
    xml = xml.replace('18/04/2026', d['fecha_emision'])
    xml = re.sub(r'(<w:t>)757(</w:t>)', lambda m: m.group(1)+str(d['total_cajas'])+m.group(2), xml)
    xml = re.sub(r'(<w:t>)13290(</w:t>)', lambda m: m.group(1)+str(int(float(d['peso_neto'] or 0)))+m.group(2), xml)
    fechas_prod = _listar_fechas_produccion(piqueo)
    if fechas_prod:
        viejo_lote = re.search(r'21/03/2026,.*?13/04/2026', xml, re.DOTALL)
        if viejo_lote:
            xml = xml[:viejo_lote.start()] + ', '.join(fechas_prod) + xml[viejo_lote.end():]
    cortes = listar_cortes_para_anexo(piqueo, 'espana')
    if cortes:
        trs = get_trs(xml)
        xml = reemplazar_bloque_cortes(xml, trs, 37, 57, cortes, d['total_cajas'], int(float(d['peso_neto'] or 0)))
    xml = _set_condicion_transporte(xml, _es_congelado(piqueo))
    return xml, alertas

def _gen_grecia_directo(xml, d, piqueo):
    alertas = []
    xml = xml.replace('SAN RAPHAEL MAERSK', d['vessel'])
    xml = xml.replace('MNBU0073802', d['contenedor'])
    xml = xml.replace('BAH66435', d['precinto_afip'])
    xml = xml.replace('268800865', d['bl'])
    xml = xml.replace('23192.00 KGS / 24051.11', f"{float(d['peso_neto']):.2f} KGS / {float(d['peso_bruto']):.2f}")
    xml = _reemplazar_fecha_roja_i14(xml, d['fecha_salida'])
    alertas.append('GRECIA DIRECTO: verificar meses de produccion en I.27')
    xml = xml.replace('28/07/2025', d['faena_desde'])
    xml = xml.replace('24/03/2026', d['faena_hasta'])
    xml = xml.replace('11/04/2026', d['fecha_emision'])
    xml = re.sub(r'(<w:t>)852(</w:t>)', lambda m: m.group(1)+str(d['total_cajas'])+m.group(2), xml)
    xml = re.sub(r'(<w:t>)23192(</w:t>)', lambda m: m.group(1)+str(int(float(d['peso_neto'] or 0)))+m.group(2), xml)
    fechas_faena = _listar_fechas_faena(piqueo)
    if fechas_faena:
        viejo = re.search(r'30/07/2025,.*?26/03/2026', xml, re.DOTALL)
        if viejo:
            xml = xml[:viejo.start()] + ', '.join(fechas_faena) + xml[viejo.end():]
    cortes = listar_cortes_para_anexo(piqueo, 'grecia_directo')
    if cortes:
        trs = get_trs(xml)
        xml = reemplazar_bloque_cortes(xml, trs, 38, 40, cortes, d['total_cajas'], int(float(d['peso_neto'] or 0)))
    xml = _set_condicion_transporte(xml, _es_congelado(piqueo))
    return xml, alertas

def _gen_grecia_rotterdam(xml, d, piqueo):
    alertas = []
    xml = xml.replace('SAN MARCO MAERSK', d['vessel'])
    xml = xml.replace('MMAU1106542', d['contenedor'])
    xml = xml.replace('BAD34926', d['precinto_afip'])
    xml = xml.replace('257800605', d['bl'])
    xml = xml.replace('18721.00 KGS  / 20017.91 KGS', peso_fmt_std(d['peso_neto'], d['peso_bruto']))
    xml = _reemplazar_fecha_roja_i14(xml, d['fecha_salida'])
    xml = xml.replace('26/07/2025', d['prod_desde'])
    xml = xml.replace('05/08/2025', d['prod_hasta'])
    xml = xml.replace('21/07/2025', d['faena_desde'])
    xml = xml.replace('01/08/2025', d['faena_hasta'])
    xml = xml.replace('23/08/2025', d['fecha_emision'])
    xml = re.sub(r'(<w:t>)1158(</w:t>)', lambda m: m.group(1)+str(d['total_cajas'])+m.group(2), xml)
    xml = re.sub(r'(<w:t>)18721(</w:t>)', lambda m: m.group(1)+str(int(float(d['peso_neto'] or 0)))+m.group(2), xml)
    fechas_faena = _listar_fechas_faena(piqueo)
    if fechas_faena:
        viejo = re.search(r'29/01/2025 - 30/01/2025.*?14/02/2025', xml, re.DOTALL)
        if viejo:
            xml = xml[:viejo.start()] + ' - '.join(fechas_faena) + xml[viejo.end():]
    cortes = listar_cortes_para_anexo(piqueo, 'grecia_rotterdam')
    if cortes:
        trs = get_trs(xml)
        xml = reemplazar_bloque_cortes(xml, trs, 38, 62, cortes, d['total_cajas'], int(float(d['peso_neto'] or 0)))
    xml = _set_condicion_transporte(xml, _es_congelado(piqueo))
    return xml, alertas


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
